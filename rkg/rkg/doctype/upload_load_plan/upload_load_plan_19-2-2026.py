# -*- coding: utf-8 -*-
import frappe
import csv
import io
from frappe.model.document import Document
from frappe.utils import now_datetime
from datetime import datetime
from openpyxl import load_workbook


EXPECTED_HEADERS = [
    "Load Reference No",
    "Dispatch Plan Date",
    "Payment Plan Date",
    "Model",
    "Model Name",
    "Type",
    "Variant",
    "Color",
    "Group Color",
    "Option",
    "Quantity",
]


class UploadLoadPlan(Document):

    def autoname(self):
        dt = now_datetime()
        self.name = f"ALP-{dt.strftime('%Y%m%d-%H%M%S')}"

    def on_update(self):
        if not self.attach_load_plan:
            return

        if getattr(self, "has_processed", 0):
            return

        self.process_and_insert_file()

    # =====================================================
    # MAIN ENTRY
    # =====================================================
    def process_and_insert_file(self):

        file_doc = frappe.get_doc("File", {"file_url": self.attach_load_plan})
        file_name = file_doc.file_name.lower()

        if file_name.endswith(".csv"):
            headers, rows = self.read_csv(file_doc)
        elif file_name.endswith(".xlsx"):
            headers, rows = self.read_excel(file_doc)
        else:
            frappe.throw("Only CSV or Excel (.xlsx) files are allowed")

        # ✅ HEADER VALIDATION (SPACE SAFE + ORDER STRICT)
        self.validate_headers(headers)

        # ✅ ROW VALIDATION + PROCESS
        self.validate_and_process_rows(rows)

    # =====================================================
    # CSV READER
    # =====================================================
    def read_csv(self, file_doc):
        content = file_doc.get_content()
        reader = csv.reader(io.StringIO(content))

        raw_headers = next(reader)
        headers = [h.strip() for h in raw_headers]

        rows = []
        for row in reader:
            rows.append(dict(zip(headers, [str(v).strip() for v in row])))

        return headers, rows

    # =====================================================
    # EXCEL READER
    # =====================================================
    def read_excel(self, file_doc):
        wb = load_workbook(file_doc.get_full_path(), data_only=True)
        ws = wb.active

        raw_headers = [str(c.value).strip() for c in ws[1]]
        headers = raw_headers[:]

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, value in enumerate(row):
                row_dict[headers[i]] = str(value).strip() if value is not None else ""
            rows.append(row_dict)

        return headers, rows

    # =====================================================
    # HEADER VALIDATION
    # =====================================================
    def validate_headers(self, headers):

        expected = [h.strip() for h in EXPECTED_HEADERS]
        incoming = [h.strip() for h in headers]

        if incoming != expected:
            error_msg = (
                "❌ Header validation failed!\n\n"
                "Expected column order:\n"
                f"{expected}\n\n"
                "Found column order:\n"
                f"{incoming}\n\n"
                "✔ Column names must match exactly\n"
                "✔ Order must be same\n"
                "✔ Extra / missing columns are not allowed"
            )
            frappe.throw(error_msg)

    # =====================================================
    # ROW VALIDATION + PROCESSING
    # =====================================================
    def validate_and_process_rows(self, rows):

        error_rows = []
        valid_rows = []

        for idx, row in enumerate(rows, start=2):

            errors = []

            # Required fields
            ref = row.get("Load Reference No")
            if not ref:
                errors.append("Load Reference No missing")

            qty = row.get("Quantity")
            try:
                qty = int(qty)
                if qty <= 0:
                    errors.append("Quantity must be > 0")
            except Exception:
                errors.append("Invalid Quantity")

            dispatch_date = self.parse_date(
                row.get("Dispatch Plan Date"),
                "Dispatch Plan Date",
                errors
            )
            payment_date = self.parse_date(
                row.get("Payment Plan Date"),
                "Payment Plan Date",
                errors
            )

            for f in ["Model", "Model Name", "Type", "Variant", "Color"]:
                if not row.get(f):
                    errors.append(f"{f} is required")

            if errors:
                error_rows.append(f"Row {idx}: " + "; ".join(errors))
                continue

            valid_rows.append({
                "load_reference_no": ref,
                "dispatch_date": dispatch_date,
                "payment_date": payment_date,
                "model": row["Model"],
                "model_name": row["Model Name"],
                "type": row["Type"],
                "variant": row["Variant"],
                "color": row["Color"],
                "group_color": row.get("Group Color"),
                "option": row.get("Option"),
                "quantity": qty,
            })

        if error_rows:
            frappe.throw(
                "❌ Validation Failed\n\n" +
                "\n".join(error_rows[:20]) +
                ("\n...more errors" if len(error_rows) > 20 else "")
            )

        self.create_or_update_load_plans(valid_rows)

    # =====================================================
    # DATE PARSER
    # =====================================================
    def parse_date(self, value, label, errors):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except Exception:
                pass
        errors.append(f"Invalid {label}")
        return None

    # =====================================================
    # CREATE / UPDATE LOAD PLANS
    # =====================================================
    def create_or_update_load_plans(self, rows):

        created, updated = 0, 0

        for r in rows:

            lp_name = frappe.db.get_value(
                "Load Plan",
                {"load_reference_no": r["load_reference_no"], "docstatus": ["!=", 2]},
                "name"
            )

            if lp_name:
                lp = frappe.get_doc("Load Plan", lp_name)
                lp.total_qty += r["quantity"]
                updated += 1
            else:
                lp = frappe.new_doc("Load Plan")
                lp.load_reference_no = r["load_reference_no"]
                lp.dispatch_plan_date = r["dispatch_date"]
                lp.payment_plan_date = r["payment_date"]
                lp.total_qty = r["quantity"]
                lp.created_by_upload = self.name
                created += 1

            lp.append("load_items", {
                "model": r["model"],
                "model_name": r["model_name"],
                "type": r["type"],
                "variant": r["variant"],
                "color": r["color"],
                "group_color": r["group_color"],
                "option": r["option"],
                "quantity": r["quantity"],
            })

            lp.status = "Planned"
            lp.save(ignore_permissions=True)

            if lp.docstatus == 0:
                lp.submit()

        self.has_processed = 1
        self.processed_rows = len(rows)
        self.save(ignore_permissions=True)

        frappe.msgprint(
            f"✅ Upload successful\nCreated: {created}\nUpdated: {updated}",
            indicator="green",
            alert=True
        )


@frappe.whitelist()
def process_file_manual(upload_doc_name):
    doc = frappe.get_doc("Upload Load Plan", upload_doc_name)
    doc.process_and_insert_file()
    return "Processed successfully"
