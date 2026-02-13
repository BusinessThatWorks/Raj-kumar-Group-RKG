# Copyright (c) 2026
# License: see license.txt

import frappe
import csv
import io
from frappe.model.document import Document
from frappe.utils import nowdate
from openpyxl import load_workbook
from datetime import datetime


# =========================================================
# STRICT HEADER FORMAT (NO EXTRA SPACES)
# =========================================================
EXPECTED_HEADERS = [
    "Battery Brand",
    "Batery Type",
    "Sample Battery Serial No",
    "Sample Battery Charging Date",
    "Charging Date",
    "Frame No",
    "Key No",
]


class BatteryandKeyUpload(Document):

    # =========================================================
    # AUTO RUN AFTER SAVE
    # =========================================================
    def on_update(self):
        if not self.excel_file:
            return

        if getattr(self, "has_processed", 0):
            return

        self.process_file()

    # =========================================================
    # DELETE BATTERY INFO IF CHILD ROW REMOVED
    # =========================================================
    def before_save(self):

        # Skip for new document
        if self.is_new():
            return

        old_doc = self.get_doc_before_save()
        if not old_doc:
            return

        old_serials = {d.battery_serial_no for d in old_doc.upload_items if d.battery_serial_no}
        new_serials = {d.battery_serial_no for d in self.upload_items if d.battery_serial_no}

        removed_serials = old_serials - new_serials

        for serial in removed_serials:
            self.delete_battery_information(serial)

    # =========================================================
    # DELETE ALL RELATED BATTERY INFO ON PARENT DELETE
    # =========================================================
    def on_trash(self):

        for row in self.upload_items:
            if row.battery_serial_no:
                self.delete_battery_information(row.battery_serial_no)

    # =========================================================
    # SAFE BATTERY INFO DELETE
    # =========================================================
    def delete_battery_information(self, serial):

        battery_name = frappe.db.get_value(
            "Battery Information",
            {"battery_serial_no": serial},
            "name"
        )

        if battery_name:
            battery_doc = frappe.get_doc("Battery Information", battery_name)

            if battery_doc.docstatus == 1:
                battery_doc.cancel()

            battery_doc.delete(ignore_permissions=True)

    # =========================================================
    # MAIN PROCESS
    # =========================================================
    def process_file(self):

        file_doc = frappe.get_doc("File", {"file_url": self.excel_file})
        file_name = file_doc.file_name.lower()

        if file_name.endswith(".csv"):
            headers, rows = self.read_csv(file_doc)
        elif file_name.endswith(".xlsx"):
            headers, rows = self.read_excel(file_doc)
        else:
            frappe.throw("Only CSV or Excel (.xlsx) allowed")

        self.validate_headers(headers)
        self.insert_rows(rows)

    # =========================================================
    # CSV READER
    # =========================================================
    def read_csv(self, file_doc):
        content = file_doc.get_content()
        reader = csv.reader(io.StringIO(content))

        raw_headers = next(reader)
        headers = [str(h).strip() for h in raw_headers]

        rows = []
        for row in reader:
            cleaned = [str(v).strip() if v else "" for v in row]
            rows.append(dict(zip(headers, cleaned)))

        return headers, rows

    # =========================================================
    # EXCEL READER
    # =========================================================
    def read_excel(self, file_doc):
        wb = load_workbook(file_doc.get_full_path(), data_only=True)
        ws = wb.active

        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, value in enumerate(row):
                row_dict[headers[i]] = str(value).strip() if value else ""
            rows.append(row_dict)

        return headers, rows

    # =========================================================
    # HEADER VALIDATION
    # =========================================================
    def validate_headers(self, headers):

        incoming = [h.strip() for h in headers]
        expected = [h.strip() for h in EXPECTED_HEADERS]

        if incoming != expected:
            frappe.throw(
                f"""❌ Header Format Incorrect

Expected:
{expected}

Found:
{incoming}
"""
            )

    # =========================================================
    # INSERT CHILD + CREATE BATTERY INFORMATION
    # =========================================================
    def insert_rows(self, rows):

        self.set("upload_items", [])
        error_rows = []
        inserted = 0

        for idx, row in enumerate(rows, start=2):

            errors = []

            frame_no_value = row.get("Frame No", "").strip()
            battery_serial = row.get("Sample Battery Serial No", "").strip()

            if not frame_no_value:
                errors.append("Frame No missing")

            if not battery_serial:
                errors.append("Battery Serial No missing")

            # Validate Item
            item_name = frappe.db.get_value(
                "Item",
                {"item_code": frame_no_value},
                "name"
            )

            if not item_name:
                errors.append(f"Item with item_code '{frame_no_value}' not found")

            charging_date = self.parse_date(
                row.get("Charging Date", "").strip(),
                "Charging Date",
                errors
            )

            battery_charging_date = row.get("Sample Battery Charging Date", "").strip()

            if errors:
                error_rows.append(f"Row {idx}: " + "; ".join(errors))
                continue

            # 1️⃣ Insert child row
            self.append("upload_items", {
                "frame_no": item_name,
                "item_code": frame_no_value,
                "battery_brand": row.get("Battery Brand", "").strip(),
                "battery_type": row.get("Batery Type", "").strip(),
                "battery_serial_no": battery_serial,
                "battery_charging_date": battery_charging_date,
                "charging_date": charging_date,
                "key_no": row.get("Key No", "").strip(),
            })

            # 2️⃣ Create + Submit Battery Information
            if not frappe.db.exists("Battery Information", {
                "battery_serial_no": battery_serial
            }):

                battery_doc = frappe.get_doc({
                    "doctype": "Battery Information",
                    "battery_serial_no": battery_serial,
                    "battery_brand": row.get("Battery Brand", "").strip(),
                    "battery_type": row.get("Batery Type", "").strip(),
                    "charging_date": charging_date,
                    "battery_charging_date": battery_charging_date,
                })

                battery_doc.insert(ignore_permissions=True)

                if battery_doc.docstatus == 0:
                    battery_doc.submit()

            inserted += 1

        if error_rows:
            frappe.throw(
                "❌ Validation Failed\n\n" +
                "\n".join(error_rows[:20]) +
                ("\n...more errors" if len(error_rows) > 20 else "")
            )

        self.date = nowdate()
        self.has_processed = 1
        # self.save(ignore_permissions=True)

        frappe.msgprint(
            f"✅ Upload Successful\nInserted Rows: {inserted}",
            indicator="green",
            alert=True
        )

    # =========================================================
    # DATE PARSER
    # =========================================================
    def parse_date(self, value, label, errors):

        if not value:
            return None

        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except Exception:
                pass

        errors.append(f"Invalid {label}")
        return None


# =========================================================
# MANUAL BUTTON
# =========================================================
@frappe.whitelist()
def process_battery_key_upload(name):
    doc = frappe.get_doc("Battery and Key Upload", name)
    doc.process_file()
    return "Processed Successfully"
